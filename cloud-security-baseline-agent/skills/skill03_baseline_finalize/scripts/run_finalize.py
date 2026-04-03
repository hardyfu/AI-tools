import argparse
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from runtime.project_root import resolve_project_root

PROJECT_ROOT = resolve_project_root(__file__)
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from runtime.ollama_runtime import build_azure_openai_runtime
from runtime.text_utils import clean_statement_noise, load_json, write_json


DOMAIN_TITLES = {
    "identity": "Identity and Access Management",
    "logging": "Logging and Monitoring",
    "network": "Network Security",
    "encryption": "Encryption and Key Management",
    "data protection": "Data Protection",
    "general": "General Platform Security",
}

SHEET_ORDER = [
    "Summary",
    "Document Sections",
    "Standard Coverage",
    "Benchmark Extensions",
    "Organization Specific",
    "Baseline Candidates",
    "Recommendations CN",
]

SECTION_FIELDS = ["title", "purpose", "scope", "principles", "governance", "domain_overviews"]


def _bucket(mapping: list[dict[str, Any]], decision: str) -> list[dict[str, Any]]:
    return [item for item in mapping if item.get("decision") == decision]


def _title(category: str) -> str:
    return DOMAIN_TITLES.get(category, "Other Controls")


def _render_trace(item: dict[str, Any]) -> str:
    trace: list[str] = []
    matched = item.get("matched_global_policy_requirement")
    benchmark = item.get("third_party_requirement", {})
    if matched:
        trace.append(str(matched.get("source_requirement_id") or matched.get("requirement_id") or "").strip())
    trace.append(str(benchmark.get("source_requirement_id") or benchmark.get("requirement_id") or "").strip())
    trace = [value for value in trace if value]
    return ", ".join(trace)


def _normalize_text_list(value: Any) -> list[str]:
    if isinstance(value, str):
        cleaned = clean_statement_noise(value)
        return [cleaned] if cleaned else []
    if not isinstance(value, list):
        return []
    normalized: list[str] = []
    char_buffer: list[str] = []
    for item in value:
        if not isinstance(item, str):
            continue
        candidate = clean_statement_noise(item)
        if not candidate:
            continue
        if len(candidate) == 1:
            char_buffer.append(candidate)
            continue
        if char_buffer:
            merged = clean_statement_noise("".join(char_buffer))
            if merged:
                normalized.append(merged)
            char_buffer = []
        normalized.append(candidate)
    if char_buffer:
        merged = clean_statement_noise("".join(char_buffer))
        if merged:
            normalized.append(merged)
    return normalized


def _join_lines(lines: list[str]) -> str:
    return "\n".join(line for line in lines if line).strip()


def _group_active_controls(analysis: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in analysis.get("baseline_mapping", []):
        if item.get("decision") == "gap":
            continue
        category = item.get("third_party_requirement", {}).get("category", "general")
        grouped.setdefault(category, []).append(item)
    return grouped


def _build_fallback_final_payload(
    case_name: str,
    analysis: dict[str, Any],
    standard_coverage: dict[str, Any],
    benchmark_extensions: dict[str, Any],
    baseline_candidates: dict[str, Any],
) -> dict[str, Any]:
    coverage_summary = standard_coverage.get("summary", {})
    extension_summary = benchmark_extensions.get("summary", {})
    candidate_summary = baseline_candidates.get("summary", {})
    grouped = _group_active_controls(analysis)
    domain_overviews: list[dict[str, str]] = []
    for domain in sorted(grouped):
        items = grouped[domain]
        matched_count = sum(1 for item in items if item.get("decision") == "aligned")
        partial_count = sum(1 for item in items if item.get("decision") == "partial")
        gap_count = sum(1 for item in items if item.get("decision") == "gap")
        domain_overviews.append(
            {
                "domain": _title(domain),
                "summary": (
                    f"This domain contains {len(items)} active control mappings, "
                    f"including {matched_count} aligned, {partial_count} partial, and {gap_count} pending items."
                ),
            }
        )
    if not domain_overviews:
        domain_overviews = [
            {
                "domain": "General Platform Security",
                "summary": "The baseline consolidates cloud security controls into a review-ready production workbook.",
            }
        ]
    return {
        "title": "Alibaba Cloud Security Baseline Final",
        "purpose": [
            f"Establish a review-ready Alibaba Cloud baseline for case {case_name}.",
            f"Summarize {coverage_summary.get('covered', 0)} covered requirements, {coverage_summary.get('partially_covered', 0)} partially covered requirements, and {coverage_summary.get('not_addressed_by_benchmark', 0)} uncovered requirements.",
        ],
        "scope": [
            "Include the Global Policy and Third-Party Standard inputs used for this case.",
            f"Reflect {extension_summary.get('total_extensions', 0)} benchmark extensions and {candidate_summary.get('total_candidates', 0)} baseline candidates.",
        ],
        "principles": [
            "Keep traceability to both source documents.",
            "Treat benchmark gaps conservatively and preserve organization-specific requirements.",
            "Render control content locally from the approved structured analysis.",
        ],
        "governance": [
            "Review exceptions and pending controls with the relevant platform owners.",
            "Keep the workbook schema stable for downstream review and operations.",
        ],
        "domain_overviews": domain_overviews,
    }


def _validate_final_payload(payload: dict[str, Any]) -> dict[str, Any]:
    missing = [field for field in SECTION_FIELDS if field not in payload]
    if missing:
        raise RuntimeError(f"LLM finalization output missing fields: {', '.join(missing)}")

    title = clean_statement_noise(str(payload.get("title", "")).strip()) or "Alibaba Cloud Security Baseline Final"
    normalized = {
        "title": title,
        "purpose": _normalize_text_list(payload.get("purpose", [])),
        "scope": _normalize_text_list(payload.get("scope", [])),
        "principles": _normalize_text_list(payload.get("principles", [])),
        "governance": _normalize_text_list(payload.get("governance", [])),
        "domain_overviews": [],
    }
    if not all(normalized[key] for key in ("purpose", "scope", "principles", "governance")):
        raise RuntimeError("LLM finalization output must provide non-empty purpose, scope, principles, and governance sections.")

    domain_overviews = payload.get("domain_overviews", [])
    if not isinstance(domain_overviews, list) or not domain_overviews:
        raise RuntimeError("LLM finalization output must provide at least one domain overview.")
    for item in domain_overviews:
        if not isinstance(item, dict):
            continue
        domain = clean_statement_noise(str(item.get("domain", "")).strip())
        summary = clean_statement_noise(str(item.get("summary", "")).strip())
        if domain and summary:
            normalized["domain_overviews"].append({"domain": domain, "summary": summary})
    if not normalized["domain_overviews"]:
        raise RuntimeError("LLM finalization output did not contain usable domain_overviews.")
    return normalized


def _set_sheet_style(sheet: Any, widths: list[tuple[str, float]]) -> None:
    header_fill = PatternFill("solid", start_color="D9EAF7", end_color="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in widths:
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"


def _markdown_lines_to_rows(text: str) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    current_section = "General"
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("#"):
            current_section = clean_statement_noise(line.lstrip("#").strip()) or current_section
            continue
        if line.startswith("-"):
            content = clean_statement_noise(line[1:].strip())
        else:
            content = clean_statement_noise(line)
        if content:
            rows.append((current_section, content))
    return rows


def _build_workbook(
    output_path: Path,
    *,
    case_name: str,
    final_payload: dict[str, Any],
    analysis: dict[str, Any],
    standard_coverage: dict[str, Any],
    benchmark_extensions: dict[str, Any],
    baseline_candidates: dict[str, Any],
    recommendations_text: str,
    finalization_model: str,
    runtime_status: dict[str, Any],
    used_llm: bool,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Field", "Value"])
    summary_sheet_rows = [
        ("case_name", case_name),
        ("document_title", final_payload["title"]),
        ("target_platform", "Alibaba Cloud"),
        ("finalization_model", finalization_model),
        ("covered_count", standard_coverage["summary"]["covered"]),
        ("partially_covered_count", standard_coverage["summary"]["partially_covered"]),
        ("not_addressed_by_benchmark_count", standard_coverage["summary"]["not_addressed_by_benchmark"]),
        ("organization_specific_count", standard_coverage["summary"]["organization_specific"]),
        ("benchmark_extension_count", benchmark_extensions["summary"]["total_extensions"]),
        ("baseline_candidate_count", baseline_candidates["summary"]["total_candidates"]),
        ("llm_used", str(bool(used_llm))),
        ("provider", runtime_status.get("provider", "unknown")),
    ]
    for row in summary_sheet_rows:
        summary_sheet.append(list(row))
    _set_sheet_style(summary_sheet, [("A", 28), ("B", 90)])

    sections_sheet = workbook.create_sheet("Document Sections")
    sections_sheet.append(["Section", "Content"])
    sections_sheet.append(["Title", final_payload["title"]])
    sections_sheet.append(["Purpose", _join_lines(final_payload["purpose"])])
    sections_sheet.append(["Scope", _join_lines(final_payload["scope"])])
    sections_sheet.append(["Principles", _join_lines(final_payload["principles"])])
    sections_sheet.append(["Governance", _join_lines(final_payload["governance"])])
    for item in final_payload["domain_overviews"]:
        sections_sheet.append([f"Domain Overview: {item['domain']}", item["summary"]])
    _set_sheet_style(sections_sheet, [("A", 30), ("B", 120)])

    coverage_sheet = workbook.create_sheet("Standard Coverage")
    coverage_sheet.append([
        "Global Requirement ID",
        "Global Source Requirement ID",
        "Global Section",
        "Global Statement",
        "Category",
        "Priority",
        "Service",
        "Coverage Status",
        "Matched Benchmark Requirements",
        "Rationale",
    ])
    for item in standard_coverage.get("rows", []):
        matched_label = "; ".join(
            f"{row.get('source_requirement_id') or row.get('requirement_id')}: {clean_statement_noise(str(row.get('statement', '')))}"
            for row in item.get("matched_benchmark_requirements", [])
        )
        coverage_sheet.append([
            item.get("global_requirement_id", ""),
            item.get("global_source_requirement_id", ""),
            item.get("global_section", ""),
            clean_statement_noise(str(item.get("global_statement", ""))),
            item.get("category", ""),
            item.get("priority", ""),
            item.get("service", ""),
            item.get("coverage_status", ""),
            matched_label,
            clean_statement_noise(str(item.get("coverage_rationale", ""))),
        ])
    _set_sheet_style(coverage_sheet, [("A", 18), ("B", 22), ("C", 22), ("D", 60), ("E", 18), ("F", 12), ("G", 16), ("H", 24), ("I", 72), ("J", 54)])

    extension_sheet = workbook.create_sheet("Benchmark Extensions")
    extension_sheet.append([
        "Domain",
        "Extension Type",
        "Decision",
        "Benchmark Requirement ID",
        "Benchmark Source Requirement ID",
        "Benchmark Section",
        "Benchmark Statement",
        "Benchmark Category",
        "Benchmark Service",
        "Related Global Requirements",
        "Baseline Candidate",
        "Candidate Priority",
        "Rationale",
    ])
    for item in benchmark_extensions.get("rows", []):
        related_label = "; ".join(
            f"{row.get('global_source_requirement_id') or row.get('global_requirement_id')}: {clean_statement_noise(str(row.get('global_statement', '')))}"
            for row in item.get("related_global_requirements", [])
        )
        extension_sheet.append([
            _title(str(item.get("category", "general"))),
            item.get("extension_type", ""),
            item.get("decision", ""),
            item.get("benchmark_requirement_id", ""),
            item.get("benchmark_source_requirement_id", ""),
            item.get("benchmark_section", ""),
            clean_statement_noise(str(item.get("benchmark_statement", ""))),
            item.get("category", ""),
            item.get("service", ""),
            related_label,
            str(bool(item.get("baseline_candidate", False))),
            item.get("candidate_priority", ""),
            clean_statement_noise(str(item.get("extension_rationale", ""))),
        ])
    _set_sheet_style(extension_sheet, [("A", 26), ("B", 24), ("C", 16), ("D", 18), ("E", 22), ("F", 24), ("G", 56), ("H", 18), ("I", 18), ("J", 70), ("K", 18), ("L", 16), ("M", 48)])

    org_only_sheet = workbook.create_sheet("Organization Specific")
    org_only_sheet.append([
        "Global Requirement ID",
        "Global Source Requirement ID",
        "Global Section",
        "Global Statement",
        "Category",
        "Priority",
        "Rationale",
    ])
    for item in standard_coverage.get("rows", []):
        if item.get("coverage_status") != "organization_specific":
            continue
        org_only_sheet.append([
            item.get("global_requirement_id", ""),
            item.get("global_source_requirement_id", ""),
            item.get("global_section", ""),
            clean_statement_noise(str(item.get("global_statement", ""))),
            item.get("category", ""),
            item.get("priority", ""),
            clean_statement_noise(str(item.get("coverage_rationale", ""))),
        ])
    _set_sheet_style(org_only_sheet, [("A", 18), ("B", 22), ("C", 24), ("D", 68), ("E", 18), ("F", 12), ("G", 48)])

    candidates_sheet = workbook.create_sheet("Baseline Candidates")
    candidates_sheet.append([
        "Candidate ID",
        "Source Benchmark Requirement ID",
        "Source Benchmark Source Requirement ID",
        "Proposed Control Title",
        "Proposed Control Statement",
        "Category",
        "Priority",
        "Service",
        "Extension Type",
        "Related Global Requirements",
        "Rationale",
    ])
    for item in baseline_candidates.get("rows", []):
        related_label = "; ".join(
            f"{row.get('global_source_requirement_id') or row.get('global_requirement_id')}: {clean_statement_noise(str(row.get('global_statement', '')))}"
            for row in item.get("related_global_requirements", [])
        )
        candidates_sheet.append([
            item.get("candidate_id", ""),
            item.get("source_benchmark_requirement_id", ""),
            item.get("source_benchmark_source_requirement_id", ""),
            clean_statement_noise(str(item.get("proposed_control_title", ""))),
            clean_statement_noise(str(item.get("proposed_control_statement", ""))),
            item.get("category", ""),
            item.get("candidate_priority", ""),
            item.get("service", ""),
            item.get("extension_type", ""),
            related_label,
            clean_statement_noise(str(item.get("reason_for_inclusion", ""))),
        ])
    _set_sheet_style(candidates_sheet, [("A", 16), ("B", 18), ("C", 22), ("D", 42), ("E", 60), ("F", 18), ("G", 12), ("H", 16), ("I", 22), ("J", 62), ("K", 52)])

    rec_sheet = workbook.create_sheet("Recommendations CN")
    rec_sheet.append(["Section", "Content"])
    for row in _markdown_lines_to_rows(recommendations_text):
        rec_sheet.append(list(row))
    _set_sheet_style(rec_sheet, [("A", 30), ("B", 110)])

    if workbook.sheetnames != SHEET_ORDER:
        raise RuntimeError(f"Workbook schema mismatch. Expected sheets: {', '.join(SHEET_ORDER)}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def validate_workbook(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing final workbook: {path}")
    workbook = load_workbook(path, read_only=True)
    if workbook.sheetnames != SHEET_ORDER:
        raise RuntimeError(f"Workbook schema mismatch. Expected sheets: {', '.join(SHEET_ORDER)}, got {', '.join(workbook.sheetnames)}")
    expected_headers = {
        "Summary": ["Field", "Value"],
        "Document Sections": ["Section", "Content"],
        "Standard Coverage": [
            "Global Requirement ID",
            "Global Source Requirement ID",
            "Global Section",
            "Global Statement",
            "Category",
            "Priority",
            "Service",
            "Coverage Status",
            "Matched Benchmark Requirements",
            "Rationale",
        ],
        "Benchmark Extensions": [
            "Domain",
            "Extension Type",
            "Decision",
            "Benchmark Requirement ID",
            "Benchmark Source Requirement ID",
            "Benchmark Section",
            "Benchmark Statement",
            "Benchmark Category",
            "Benchmark Service",
            "Related Global Requirements",
            "Baseline Candidate",
            "Candidate Priority",
            "Rationale",
        ],
        "Organization Specific": [
            "Global Requirement ID",
            "Global Source Requirement ID",
            "Global Section",
            "Global Statement",
            "Category",
            "Priority",
            "Rationale",
        ],
        "Baseline Candidates": [
            "Candidate ID",
            "Source Benchmark Requirement ID",
            "Source Benchmark Source Requirement ID",
            "Proposed Control Title",
            "Proposed Control Statement",
            "Category",
            "Priority",
            "Service",
            "Extension Type",
            "Related Global Requirements",
            "Rationale",
        ],
        "Recommendations CN": ["Section", "Content"],
    }
    for sheet_name, headers in expected_headers.items():
        sheet = workbook[sheet_name]
        actual = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        if actual != headers:
            raise RuntimeError(f"Workbook sheet {sheet_name} header mismatch. Expected {headers}, got {actual}")


def run(case_name: str) -> tuple[Path, Path]:
    working_dir = PROJECT_ROOT / "cases" / case_name / "working"
    profile_path = working_dir / "project_profile.json"
    analysis_path = working_dir / "baseline_analysis.json"
    standard_coverage_path = working_dir / "standard_coverage.json"
    benchmark_extensions_path = working_dir / "benchmark_extensions.json"
    baseline_candidates_path = working_dir / "baseline_candidates.json"
    controls_path = working_dir / "baseline_controls.md"
    report_path = working_dir / "baseline_report.md"
    recommendations_path = working_dir / "baseline_priority_recommendations_cn.md"
    for required in (
        profile_path,
        analysis_path,
        standard_coverage_path,
        benchmark_extensions_path,
        baseline_candidates_path,
        controls_path,
        report_path,
        recommendations_path,
    ):
        if not required.exists():
            raise FileNotFoundError(f"Missing skill03 input artifact: {required}")

    profile = load_json(profile_path)
    analysis = load_json(analysis_path)
    standard_coverage = load_json(standard_coverage_path)
    benchmark_extensions = load_json(benchmark_extensions_path)
    baseline_candidates = load_json(baseline_candidates_path)
    controls_text = controls_path.read_text(encoding="utf-8")
    report_text = report_path.read_text(encoding="utf-8")
    recommendations_text = recommendations_path.read_text(encoding="utf-8")

    runtime, runtime_status = build_azure_openai_runtime(profile)
    finalization_model = str(
        runtime_status.get("finalization_model")
        or getattr(getattr(runtime, "config", None), "finalization_model", "")
        or "gpt-5.4"
    )

    used_llm = False
    fallback_reason = ""
    if runtime is None:
        fallback_reason = f"Azure OpenAI runtime unavailable: {runtime_status.get('skip_reason')}"
        debug_path = working_dir / "skill03_debug.json"
        write_json(
            debug_path,
            {
                "case_name": case_name,
                "provider": runtime_status.get("provider", "unknown"),
                "available": bool(runtime_status.get("available", False)),
                "finalization_model": finalization_model,
                "used_llm": False,
                "fallback_reason": fallback_reason,
                "source_artifacts": {
                    "baseline_analysis": str(analysis_path.relative_to(PROJECT_ROOT)),
                    "standard_coverage": str(standard_coverage_path.relative_to(PROJECT_ROOT)),
                    "benchmark_extensions": str(benchmark_extensions_path.relative_to(PROJECT_ROOT)),
                    "baseline_candidates": str(baseline_candidates_path.relative_to(PROJECT_ROOT)),
                    "baseline_controls": str(controls_path.relative_to(PROJECT_ROOT)),
                    "baseline_report": str(report_path.relative_to(PROJECT_ROOT)),
                    "priority_recommendations_cn": str(recommendations_path.relative_to(PROJECT_ROOT)),
                },
            },
        )
        raise RuntimeError(f"skill03 requires LLM runtime, but it is unavailable: {runtime_status.get('skip_reason')}")

    system_prompt = (
        "You are an enterprise cloud security standards editor. "
        "Produce a formal, review-ready English baseline document structure from the approved baseline analysis. "
        "Do not redo the mapping analysis. Do not convert gap or deferred controls into covered controls. "
        "Return JSON only. Do not return Markdown or commentary."
    )
    user_prompt = (
        "Generate the JSON structure for a formal English document titled 'Alibaba Cloud Security Baseline Final'.\n"
        "Requirements:\n"
        "1. The JSON fields must be: title, purpose, scope, principles, governance, domain_overviews.\n"
        "2. domain_overviews must be an array of objects with fields: domain, summary.\n"
        "3. Write only document-level formal wording and domain summaries. Do not rewrite each control and do not emit pending control details.\n"
        "4. Control details will be rendered locally. Your job is to provide formal English section wording and domain overviews.\n"
        "5. Use concise, formal English. Avoid parser noise.\n\n"
        f"[baseline_analysis.json]\n{analysis}\n\n"
        f"[standard_coverage.json]\n{standard_coverage}\n\n"
        f"[benchmark_extensions.json]\n{benchmark_extensions}\n\n"
        f"[baseline_candidates.json]\n{baseline_candidates}\n\n"
        f"[baseline_controls.md]\n{controls_text[:12000]}\n\n"
        f"[baseline_report.md]\n{report_text[:8000]}\n\n"
        f"[baseline_priority_recommendations_cn.md]\n{recommendations_text[:8000]}"
    )

    try:
        final_payload_raw = runtime.chat_json(
            model=finalization_model,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            num_predict=2200,
        )
        final_payload = _validate_final_payload(final_payload_raw)
        used_llm = True
    except Exception as exc:
        fallback_reason = str(exc)
        final_payload = _build_fallback_final_payload(
            case_name,
            analysis,
            standard_coverage,
            benchmark_extensions,
            baseline_candidates,
        )
        debug_path = working_dir / "skill03_debug.json"
        write_json(
            debug_path,
            {
                "case_name": case_name,
                "provider": runtime_status.get("provider", "unknown"),
                "available": bool(runtime_status.get("available", False)),
                "finalization_model": finalization_model,
                "used_llm": False,
                "fallback_reason": fallback_reason,
                "final_payload_source": "fallback",
                "source_artifacts": {
                    "baseline_analysis": str(analysis_path.relative_to(PROJECT_ROOT)),
                    "standard_coverage": str(standard_coverage_path.relative_to(PROJECT_ROOT)),
                    "benchmark_extensions": str(benchmark_extensions_path.relative_to(PROJECT_ROOT)),
                    "baseline_candidates": str(baseline_candidates_path.relative_to(PROJECT_ROOT)),
                    "baseline_controls": str(controls_path.relative_to(PROJECT_ROOT)),
                    "baseline_report": str(report_path.relative_to(PROJECT_ROOT)),
                    "priority_recommendations_cn": str(recommendations_path.relative_to(PROJECT_ROOT)),
                },
            },
        )
    output_path = working_dir / "final_baseline.xlsx"
    _build_workbook(
        output_path,
        case_name=case_name,
        final_payload=final_payload,
        analysis=analysis,
        standard_coverage=standard_coverage,
        benchmark_extensions=benchmark_extensions,
        baseline_candidates=baseline_candidates,
        recommendations_text=recommendations_text,
        finalization_model=finalization_model,
        runtime_status=runtime_status,
        used_llm=used_llm,
    )

    debug_path = working_dir / "skill03_debug.json"
    write_json(
        debug_path,
        {
            "case_name": case_name,
            "provider": runtime_status.get("provider", "unknown"),
            "available": bool(runtime_status.get("available", False)),
            "finalization_model": finalization_model,
            "used_llm": used_llm,
            "fallback_reason": fallback_reason,
            "final_payload_source": "llm" if used_llm else "fallback",
            "source_artifacts": {
                "baseline_analysis": str(analysis_path.relative_to(PROJECT_ROOT)),
                "standard_coverage": str(standard_coverage_path.relative_to(PROJECT_ROOT)),
                "benchmark_extensions": str(benchmark_extensions_path.relative_to(PROJECT_ROOT)),
                "baseline_candidates": str(baseline_candidates_path.relative_to(PROJECT_ROOT)),
                "baseline_controls": str(controls_path.relative_to(PROJECT_ROOT)),
                "baseline_report": str(report_path.relative_to(PROJECT_ROOT)),
                "priority_recommendations_cn": str(recommendations_path.relative_to(PROJECT_ROOT)),
                "final_baseline_xlsx": str(output_path.relative_to(PROJECT_ROOT)),
            },
        },
    )
    return output_path, debug_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Run skill03 baseline finalization")
    parser.add_argument("--case", required=True, dest="case_name")
    args = parser.parse_args()
    try:
        outputs = run(args.case_name)
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1
    for output in outputs:
        print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

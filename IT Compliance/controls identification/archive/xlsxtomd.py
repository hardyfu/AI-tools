import openpyxl

def excel_to_markdown(file_path, sheet_name=0):
    # 加载 Excel 文件
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # 获取指定的 sheet（默认第一个）
    if isinstance(sheet_name, int):
        sheet = wb.worksheets[sheet_name]
    else:
        sheet = wb[sheet_name]

    # 获取合并单元格的映射关系
    # merged_cells 存储了每个合并区域左上角的值
    merged_ranges = sheet.merged_cells.ranges

    def get_cell_value(row, col):
        """获取单元格的值，如果是合并单元格，则返回该区域左上角的值"""
        cell = sheet.cell(row=row, column=col)
        for r in merged_ranges:
            if cell.coordinate in r:
                # 返回合并区域左上角单元格的值
                return sheet.cell(row=r.min_row, column=r.min_col).value
        return cell.value

    # 读取所有数据
    rows = []
    max_row = sheet.max_row
    max_col = sheet.max_column

    for r in range(1, max_row + 1):
        row_data = []
        for c in range(1, max_col + 1):
            val = get_cell_value(r, c)
            if val is None:
                row_data.append("")
            else:
                # 关键步骤：将单元格内的换行符 \n 替换为 HTML 的 <br>
                row_data.append(str(val).replace('\n', '<br>'))
        rows.append(row_data)

    # 转换为 Markdown 字符串
    md_output = []
    for i, row in enumerate(rows):
        # 拼接一行数据
        line = "| " + " | ".join(row) + " |"
        md_output.append(line)
        
        # 如果是表头行，添加分割线
        if i == 0:
            separator = "| " + " | ".join(["---"] * len(row)) + " |"
            md_output.append(separator)

    return "\n".join(md_output)

# 使用示例
if __name__ == "__main__":
    # 请确保已安装 openpyxl: pip install openpyxl
    file_name = "/Users/ryan/Desktop/Book1.xlsx"  # 替换为你的文件名
    try:
        markdown_text = excel_to_markdown(file_name)
        print(markdown_text)
        
        # 保存到文件
        with open("/Users/ryan/Desktop/Book1.md", "w", encoding="utf-8") as f:
            f.write(markdown_text)
    except Exception as e:
        print(f"处理出错: {e}")
import json
import re
import argparse
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import openpyxl
import requests
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent
DOCX_PATH = BASE_DIR / "sample.docx"
OUTPUT_PATH = BASE_DIR / "sample_control_eval_report.json"
OUTPUT_MD_PATH = BASE_DIR / "sample_control_eval_report.md"
OLLAMA_API_URL = "http://localhost:11434/api/chat"
VISION_MODEL = "qwen3-vl:8b"
JUDGE_MODEL = "qwen3:8b"

NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "o": "urn:schemas-microsoft-com:office:office",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def log_step(message):
    print(f"[*] {message}")


def call_ollama(payload):
    response = requests.post(OLLAMA_API_URL, json=payload, timeout=180)
    response.raise_for_status()
    return response.json().get("message", {}).get("content", "")


def parse_json_text(text):
    s = (text or "").strip()
    if s.startswith("```"):
        s = s.strip("`")
        if s.lower().startswith("json"):
            s = s[4:].strip()
    return json.loads(s)


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def escape_md(text):
    return normalize_text(text).replace("\n", "<br>").replace("|", "\\|")


def is_mandatory(control):
    t = normalize_text((control or {}).get("类型"))
    return "强制" in t


def has_effective_evidence(evidence_parts):
    if not evidence_parts:
        return False
    text = "\n".join(normalize_text(x) for x in evidence_parts).strip()
    if not text:
        return False
    return text != "[图片OCR] [跳过OCR，未调用视觉模型]"


def build_combined_evidence_text(evidence_parts):
    grouped = {
        "文本": [],
        "表格": [],
        "图片OCR": [],
        "内嵌Excel": [],
        "内嵌PDF": [],
        "PDF": [],
        "嵌入对象": [],
        "其他": [],
    }
    for part in evidence_parts or []:
        text = normalize_text(part)
        m = re.match(r"^\[([^\]]+)\]\s*(.*)$", text, re.S)
        if not m:
            grouped["其他"].append(text)
            continue
        tag = m.group(1).strip()
        body = m.group(2).strip()
        if tag in grouped:
            grouped[tag].append(body)
        else:
            grouped["其他"].append(text)

    lines = []
    lines.append("【证据组合概览】")
    lines.append(
        "；".join(
            [
                f"文本{len(grouped['文本'])}条",
                f"表格{len(grouped['表格'])}条",
                f"图片OCR{len(grouped['图片OCR'])}条",
                f"内嵌Excel{len(grouped['内嵌Excel'])}条",
                f"内嵌PDF{len(grouped['内嵌PDF'])}条",
                f"PDF{len(grouped['PDF'])}条",
                f"嵌入对象{len(grouped['嵌入对象'])}条",
                f"其他{len(grouped['其他'])}条",
            ]
        )
    )

    order = ["文本", "表格", "图片OCR", "内嵌Excel", "内嵌PDF", "PDF", "嵌入对象", "其他"]
    for key in order:
        items = grouped[key]
        if not items:
            continue
        lines.append("")
        lines.append(f"【{key}证据】")
        for i, item in enumerate(items, start=1):
            lines.append(f"({i}) {item}")
    return "\n".join(lines).strip()


def build_mandatory_coverage(results, control_files):
    section_map = {normalize_text(r.get("heading")): r for r in results}
    missing = []
    total_mandatory = 0
    provided_mandatory = 0

    for cf in sorted(control_files):
        cmap = load_control_map(Path(cf))
        for code, cp in sorted(cmap.items(), key=lambda x: x[0]):
            if not is_mandatory(cp):
                continue
            total_mandatory += 1
            item = section_map.get(code)
            provided = bool(item and item.get("has_evidence"))
            if provided:
                provided_mandatory += 1
            else:
                missing.append({
                    "heading": code,
                    "control_file": cf,
                    "控制目标": cp.get("控制目标"),
                    "控制点": cp.get("控制点"),
                    "类型": cp.get("类型"),
                })

    return {
        "all_mandatory_have_evidence": (total_mandatory == provided_mandatory),
        "total_mandatory": total_mandatory,
        "provided_mandatory": provided_mandatory,
        "missing_mandatory": missing,
    }


def build_markdown_report(report):
    lines = []
    summary = report.get("summary", {})
    coverage = report.get("mandatory_coverage", {})
    pass_n = int(summary.get("pass", 0) or 0)
    fail_n = int(summary.get("fail", 0) or 0)
    review_n = int(summary.get("review", 0) or 0)
    provided_n = int(coverage.get("provided_mandatory", 0) or 0)
    total_mandatory_n = int(coverage.get("total_mandatory", 0) or 0)
    missing_n = max(total_mandatory_n - provided_n, 0)
    lines.append("# 控制点评估结果")
    lines.append("")
    lines.append("## 1. 执行摘要")
    lines.append("")
    lines.append(f"- 文档: `{report.get('docx_path', '')}`")
    lines.append(f"- 总标题: `{report.get('total_sections', 0)}`")
    lines.append(f"- Pass: `{summary.get('pass', 0)}`")
    lines.append(f"- Fail: `{summary.get('fail', 0)}`")
    lines.append(f"- Review: `{summary.get('review', 0)}`")
    lines.append("")
    lines.append("### 评估状态分布图")
    lines.append("")
    lines.append("```mermaid")
    lines.append("pie showData")
    lines.append('    title 单项评估状态分布')
    lines.append(f'    "Pass" : {pass_n}')
    lines.append(f'    "Fail" : {fail_n}')
    lines.append(f'    "Review" : {review_n}')
    lines.append("```")
    lines.append("")
    lines.append("## 2. 强制项证据覆盖检查")
    lines.append("")
    lines.append(
        f"- 结论: `{'通过' if coverage.get('all_mandatory_have_evidence') else '未通过'}`"
    )
    lines.append(f"- 强制项总数: `{coverage.get('total_mandatory', 0)}`")
    lines.append(f"- 已提供证据: `{coverage.get('provided_mandatory', 0)}`")
    lines.append(
        f"- 未提供证据: `{len(coverage.get('missing_mandatory', []))}`"
    )
    lines.append("")
    lines.append("### 强制项证据覆盖图")
    lines.append("")
    lines.append("```mermaid")
    lines.append("pie showData")
    lines.append('    title 强制项证据覆盖')
    lines.append(f'    "已提供证据" : {provided_n}')
    lines.append(f'    "未提供证据" : {missing_n}')
    lines.append("```")
    lines.append("")
    lines.append("未提供证据的强制项：")
    lines.append("")
    lines.append("| 编号 | 控制文件 | 控制目标 | 控制点 |")
    lines.append("| :--- | :--- | :--- | :--- |")
    missing = coverage.get("missing_mandatory", [])
    if missing:
        for m in missing:
            lines.append(
                f"| {escape_md(m.get('heading'))} | {escape_md(m.get('control_file'))} | {escape_md(m.get('控制目标'))} | {escape_md(m.get('控制点'))} |"
            )
    else:
        lines.append("| - | - | - | - |")
    lines.append("")
    lines.append("## 3. 单项评估结果")
    lines.append("")
    lines.append("| 标题 | 类型 | 是否有证据 | 状态 | 置信度 | 缺失证据 | 原因 |")
    lines.append("| :--- | :--- | :--- | :--- | :--- | :--- | :--- |")

    for r in report.get("results", []):
        judge = r.get("judge", {})
        status = judge.get("compliance_status") or r.get("status", "Review")
        confidence = judge.get("confidence", "")
        missing = judge.get("missing_evidence", "") or "-"
        reason = judge.get("reason") or r.get("reason", "")
        cp = r.get("control_point") or {}
        cp_type = cp.get("类型", "-")
        has_evidence = "是" if r.get("has_evidence") else "否"
        lines.append(
            f"| {escape_md(r.get('heading'))} | {escape_md(cp_type)} | {has_evidence} | {escape_md(status)} | {escape_md(confidence)} | {escape_md(missing)} | {escape_md(reason)} |"
        )
    return "\n".join(lines)


def parse_relationships(docx_zip):
    rel_root = ET.fromstring(docx_zip.read("word/_rels/document.xml.rels"))
    rels = {}
    for rel in rel_root.findall("pr:Relationship", NS):
        rels[rel.attrib.get("Id")] = {
            "target": rel.attrib.get("Target", ""),
            "type": rel.attrib.get("Type", ""),
        }
    return rels


def parse_style_names(docx_zip):
    root = ET.fromstring(docx_zip.read("word/styles.xml"))
    style_map = {}
    for style in root.findall("w:style", NS):
        sid = style.attrib.get(f"{{{NS['w']}}}styleId")
        name = style.find("w:name", NS)
        val = name.attrib.get(f"{{{NS['w']}}}val") if name is not None else ""
        style_map[sid] = (val or "").lower()
    return style_map


def paragraph_text(p):
    return "".join((t.text or "") for t in p.findall(".//w:t", NS)).strip()


def paragraph_style_id(p):
    ppr = p.find("w:pPr", NS)
    if ppr is None:
        return None
    pstyle = ppr.find("w:pStyle", NS)
    if pstyle is None:
        return None
    return pstyle.attrib.get(f"{{{NS['w']}}}val")


def is_heading1(style_id, style_names):
    if not style_id:
        return False
    if style_id.lower() == "heading1":
        return True
    name = style_names.get(style_id, "")
    return "heading 1" in name or "标题 1" in name or "一级标题" in name


def extract_refs_from_paragraph(p):
    refs = []
    for blip in p.findall(".//a:blip", NS):
        rid = blip.attrib.get(f"{{{NS['r']}}}embed")
        if rid:
            refs.append(rid)
    for ole in p.findall(".//o:OLEObject", NS):
        rid = ole.attrib.get(f"{{{NS['r']}}}id") or ole.attrib.get("id")
        if rid:
            refs.append(rid)
    return refs


def classify_target(target):
    low = target.lower()
    if low.endswith((".png", ".jpg", ".jpeg", ".emf", ".bmp", ".gif")):
        return "image"
    if low.endswith(".xlsx"):
        return "excel"
    if low.endswith(".pdf"):
        return "pdf"
    if low.endswith(".bin"):
        return "ole_bin"
    return "unknown"


def extract_heading_code(heading):
    """从标题中提取控制点编号，兼容如 `1-1`、`1-1 xxx`、`1-1：xxx`。"""
    text = normalize_text(heading)
    m = re.match(r"^(\d+)\s*[-－—]\s*(\d+)\b", text)
    if not m:
        return None
    return f"{int(m.group(1))}-{int(m.group(2))}"


def resolve_control_file_for_heading(heading):
    code = extract_heading_code(heading)
    if not code:
        return None
    first_no = int(code.split("-", 1)[0])
    prefix = f"{first_no:02d}."
    files = sorted(BASE_DIR.glob(f"{prefix}*.xlsx"))
    return files[0] if files else None


def load_control_map(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        code = str(row[0]).strip()
        mapping[code] = {
            "编号": code,
            "涉及领域": row[1],
            "控制目标": row[2],
            "法规要求": row[3],
            "控制点": row[4],
            "类型": row[5],
            "风险": row[6] if len(row) > 6 else None,
        }
    return mapping


def extract_excel_text(data):
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ws = wb.active
    lines = [f"[内嵌Excel] sheet={ws.title}, rows={ws.max_row}, cols={ws.max_column}"]
    max_rows = min(20, ws.max_row)
    for r in range(1, max_rows + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        vals = [normalize_text(v) for v in vals]
        lines.append(" | ".join(vals))
    return "\n".join(lines)


def extract_pdf_bytes_from_ole(data):
    start = data.find(b"%PDF")
    end = data.rfind(b"%%EOF")
    if start == -1 or end == -1 or end <= start:
        return None
    return data[start:end + 5]


def extract_pdf_text(pdf_bytes):
    reader = PdfReader(BytesIO(pdf_bytes))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        txt = page.extract_text() or ""
        txt = txt.strip()
        if txt:
            pages.append(f"[PDF p{i}] {txt}")
    return "\n".join(pages)


def ocr_image_with_vl_base64(image_bytes):
    import base64
    payload = {
        "model": VISION_MODEL,
        "messages": [
            {
                "role": "system",
                "content": "你是OCR助手。提取图片中的关键文字并简洁返回。"
            },
            {
                "role": "user",
                "content": "请识别这张证据截图里的主要文字内容。",
                "images": [base64.b64encode(image_bytes).decode("utf-8")]
            }
        ],
        "stream": False,
        "options": {"temperature": 0.1}
    }
    return normalize_text(call_ollama(payload))


def ocr_image_with_local_engine(image_bytes):
    try:
        from PIL import Image
        import pytesseract
    except Exception:
        return ""
    try:
        img = Image.open(BytesIO(image_bytes))
        text = pytesseract.image_to_string(img, lang="chi_sim+eng")
        return normalize_text(text)
    except Exception:
        return ""


def extract_image_ocr_text(image_bytes, use_vision_model=True, ocr_strategy="auto"):
    # strategy:
    # - auto: 优先函数库OCR，文本不足再调用模型
    # - local: 仅函数库OCR
    # - model: 仅模型OCR
    min_chars = 20
    local_text = ""

    if ocr_strategy in ("auto", "local"):
        local_text = ocr_image_with_local_engine(image_bytes)
        if len(local_text) >= min_chars:
            return f"[函数库OCR] {local_text}"
        if ocr_strategy == "local":
            return "[函数库OCR] 未识别到足够文本"

    if use_vision_model and ocr_strategy in ("auto", "model"):
        model_text = ocr_image_with_vl_base64(image_bytes)
        if local_text:
            return f"[函数库OCR] {local_text}\n[模型OCR] {model_text}"
        return f"[模型OCR] {model_text}"

    if local_text:
        return f"[函数库OCR] {local_text}"
    return "[图片OCR] [跳过OCR，未调用视觉模型]"


def judge_evidence(control, evidence_text):
    system_prompt = (
        "你是IT合规审计员。请根据控制点要求判断证据是否满足。"
        "只输出JSON，字段: compliance_status(Pass/Fail/Review), confidence(0-1), reason, missing_evidence。"
    )
    user_prompt = (
        f"【控制点编号】{control.get('编号')}\n"
        f"【控制目标】{normalize_text(control.get('控制目标'))}\n"
        f"【法规要求】{normalize_text(control.get('法规要求'))}\n"
        f"【控制点】{normalize_text(control.get('控制点'))}\n\n"
        f"【证据内容】\n{evidence_text[:12000]}"
    )
    payload = {
        "model": JUDGE_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "format": "json",
        "stream": False,
        "options": {"temperature": 0.1}
    }
    result = parse_json_text(call_ollama(payload))
    result["compliance_status"] = normalize_text(result.get("compliance_status")) or "Review"
    result["confidence"] = normalize_text(result.get("confidence"))
    result["reason"] = normalize_text(result.get("reason"))
    result["missing_evidence"] = normalize_text(result.get("missing_evidence"))
    return result


def build_sections_and_evidence(docx_path, use_vision_model=True, ocr_strategy="auto"):
    log_step(f"开始解析DOCX并提取证据: {docx_path}")
    sections = []
    current = None
    control_cache = {}

    with ZipFile(docx_path) as docx_zip:
        rels = parse_relationships(docx_zip)
        style_names = parse_style_names(docx_zip)
        doc_root = ET.fromstring(docx_zip.read("word/document.xml"))
        body = doc_root.find("w:body", NS)

        for block in list(body):
            tag = block.tag.rsplit("}", 1)[-1]
            if tag == "p":
                txt = paragraph_text(block)
                style_id = paragraph_style_id(block)

                if is_heading1(style_id, style_names):
                    heading = txt
                    heading_code = extract_heading_code(heading)
                    control_file = resolve_control_file_for_heading(heading)
                    control = None
                    if control_file:
                        cache_key = str(control_file)
                        if cache_key not in control_cache:
                            control_cache[cache_key] = load_control_map(control_file)
                        control = control_cache[cache_key].get(heading_code or heading)

                    current = {
                        "heading": heading,
                        "control_file": str(control_file) if control_file else None,
                        "control_point": control,
                        "evidence_parts": [],
                    }
                    sections.append(current)
                    log_step(
                        f"识别标题: {heading} -> 控制文件: {current['control_file'] or '-'}"
                    )
                    continue

                if current is None:
                    continue

                if txt:
                    current["evidence_parts"].append(f"[文本] {txt}")

                refs = extract_refs_from_paragraph(block)
                for rid in refs:
                    rel = rels.get(rid, {})
                    target = rel.get("target", "")
                    kind = classify_target(target)
                    full_name = f"word/{target}" if target and not target.startswith("word/") else target
                    if not full_name or full_name not in docx_zip.namelist():
                        continue

                    data = docx_zip.read(full_name)
                    if kind == "excel":
                        current["evidence_parts"].append(extract_excel_text(data))
                    elif kind == "image":
                        ocr_text = extract_image_ocr_text(
                            data, use_vision_model=use_vision_model, ocr_strategy=ocr_strategy
                        )
                        current["evidence_parts"].append(f"[图片OCR] {ocr_text}")
                    elif kind == "ole_bin":
                        pdf_bytes = extract_pdf_bytes_from_ole(data)
                        if pdf_bytes:
                            pdf_text = extract_pdf_text(pdf_bytes)
                            current["evidence_parts"].append(f"[内嵌PDF] {pdf_text}")
                        else:
                            current["evidence_parts"].append("[内嵌对象] 未识别为PDF")
                    elif kind == "pdf":
                        pdf_text = extract_pdf_text(data)
                        current["evidence_parts"].append(f"[PDF] {pdf_text}")
                    else:
                        current["evidence_parts"].append(f"[嵌入对象] kind={kind}, target={target}")

            elif tag == "tbl":
                if current is None:
                    continue
                rows = []
                for tr in block.findall(".//w:tr", NS):
                    vals = []
                    for tc in tr.findall(".//w:tc", NS):
                        cell = "".join((t.text or "") for t in tc.findall(".//w:t", NS)).strip()
                        vals.append(cell)
                    if vals:
                        rows.append(" | ".join(vals))
                if rows:
                    current["evidence_parts"].append("[表格]\n" + "\n".join(rows[:20]))

    log_step(f"DOCX解析完成，共识别一级标题: {len(sections)}")
    return sections


def main():
    parser = argparse.ArgumentParser(description="Sample DOCX 控制点匹配与证据判断 Demo")
    parser.add_argument("--headings", type=str, default="", help="只处理指定标题，逗号分隔，如 2-1,2-2")
    parser.add_argument("--max-sections", type=int, default=0, help="最多处理前N个标题，0表示全部")
    parser.add_argument("--extract-only", action="store_true", help="仅提取证据，不调用qwen3做合规判断")
    parser.add_argument("--no-image-ocr", action="store_true", help="不调用qwen3-vl识别图片")
    parser.add_argument(
        "--ocr-strategy",
        type=str,
        default="auto",
        choices=["auto", "local", "model"],
        help="图片OCR策略: auto(先函数库再模型) / local(仅函数库) / model(仅模型)",
    )
    parser.add_argument("--from-json", type=str, default="", help="从已有JSON结果生成Markdown，不重新调用模型")
    args = parser.parse_args()

    if args.from_json:
        log_step(f"从已有JSON生成Markdown: {args.from_json}")
        src = Path(args.from_json).resolve()
        if not src.exists():
            raise FileNotFoundError(f"找不到JSON文件: {src}")
        report = json.loads(src.read_text(encoding="utf-8"))
        OUTPUT_MD_PATH.write_text(build_markdown_report(report), encoding="utf-8")
        print(f"[+] 已输出: {OUTPUT_MD_PATH}")
        return

    if not DOCX_PATH.exists():
        raise FileNotFoundError(f"找不到文件: {DOCX_PATH}")

    if args.no_image_ocr and args.ocr_strategy == "model":
        log_step("检测到 --no-image-ocr 与 --ocr-strategy model 冲突，已自动切换为 local")
        args.ocr_strategy = "local"

    log_step(f"图片OCR策略: {args.ocr_strategy}")
    sections = build_sections_and_evidence(
        DOCX_PATH,
        use_vision_model=not args.no_image_ocr,
        ocr_strategy=args.ocr_strategy,
    )

    selected = {h.strip() for h in args.headings.split(",") if h.strip()}
    if selected:
        sections = [s for s in sections if s.get("heading") in selected]
    if args.max_sections > 0:
        sections = sections[:args.max_sections]
    log_step(f"准备处理标题数量: {len(sections)}")

    results = []
    involved_control_files = set()

    for idx, sec in enumerate(sections, start=1):
        heading = sec.get("heading")
        control = sec.get("control_point")
        evidence_text = build_combined_evidence_text(sec.get("evidence_parts", []))
        log_step(f"[{idx}/{len(sections)}] 处理标题: {heading}")

        if not control:
            results.append({
                "heading": heading,
                "control_file": sec.get("control_file"),
                "status": "Review",
                "reason": "未找到对应控制点编号",
                "evidence_preview": evidence_text[:500],
                "has_evidence": has_effective_evidence(sec.get("evidence_parts", [])),
            })
            log_step(f"[{idx}/{len(sections)}] 未匹配到控制点编号，标记为Review")
            continue

        if sec.get("control_file"):
            involved_control_files.add(sec.get("control_file"))

        if args.extract_only:
            judge = {
                "compliance_status": "Review",
                "confidence": "",
                "reason": "extract-only 模式：未调用判断模型",
                "missing_evidence": "",
            }
        else:
            try:
                log_step(f"[{idx}/{len(sections)}] 调用判断模型: {JUDGE_MODEL}")
                judge = judge_evidence(control, evidence_text)
            except Exception as e:
                judge = {
                    "compliance_status": "Review",
                    "confidence": "",
                    "reason": f"模型判断失败: {e}",
                    "missing_evidence": "",
                }
                log_step(f"[{idx}/{len(sections)}] 模型判断失败，已降级为Review")

        results.append({
            "heading": heading,
            "control_file": sec.get("control_file"),
            "control_point": control,
            "evidence_preview": evidence_text[:1200],
            "has_evidence": has_effective_evidence(sec.get("evidence_parts", [])),
            "judge": judge,
        })

    report = {
        "docx_path": str(DOCX_PATH),
        "total_sections": len(sections),
        "results": results,
        "mandatory_coverage": build_mandatory_coverage(results, involved_control_files),
        "summary": {
            "pass": sum(1 for r in results if (r.get("judge", {}).get("compliance_status") == "Pass")),
            "fail": sum(1 for r in results if (r.get("judge", {}).get("compliance_status") == "Fail")),
            "review": sum(
                1
                for r in results
                if (r.get("judge", {}).get("compliance_status") == "Review" or r.get("status") == "Review")
            ),
        },
    }

    log_step("写出JSON与Markdown报告")
    OUTPUT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    OUTPUT_MD_PATH.write_text(build_markdown_report(report), encoding="utf-8")
    print(f"[+] 已输出: {OUTPUT_PATH}")
    print(f"[+] 已输出: {OUTPUT_MD_PATH}")
    print(f"[+] 总标题: {report['total_sections']}")
    print(f"[+] Pass={report['summary']['pass']}, Fail={report['summary']['fail']}, Review={report['summary']['review']}")


if __name__ == "__main__":
    main()

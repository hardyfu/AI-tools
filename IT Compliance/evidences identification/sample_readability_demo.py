import json
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
DOCX_PATH = BASE_DIR / "sample.docx"
CONTROL_XLSX_PATH = BASE_DIR / "02.日志管理.xlsx"
OUTPUT_PATH = BASE_DIR / "sample_readability_report.json"

NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "o": "urn:schemas-microsoft-com:office:office",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def load_control_points(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = str(row[0]).strip() if row[0] is not None else ""
        if not code:
            continue
        mapping[code] = {
            "编号": code,
            "涉及领域": row[1],
            "控制目标": row[2],
            "法规要求": row[3],
            "控制点": row[4],
            "类型": row[5],
            "风险": row[6] if len(row) > 6 else None,
        }
    return mapping


def parse_relationships(docx_zip: ZipFile):
    rel_root = ET.fromstring(docx_zip.read("word/_rels/document.xml.rels"))
    rels = {}
    for rel in rel_root.findall("pr:Relationship", NS):
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target", "")
        rel_type = rel.attrib.get("Type", "")
        rels[rid] = {"target": target, "type": rel_type}
    return rels


def parse_style_names(docx_zip: ZipFile):
    style_root = ET.fromstring(docx_zip.read("word/styles.xml"))
    style_map = {}
    for style in style_root.findall("w:style", NS):
        style_id = style.attrib.get(f"{{{NS['w']}}}styleId")
        name_node = style.find("w:name", NS)
        style_name = name_node.attrib.get(f"{{{NS['w']}}}val") if name_node is not None else ""
        style_map[style_id] = (style_name or "").lower()
    return style_map


def paragraph_text(p):
    parts = []
    for t in p.findall(".//w:t", NS):
        parts.append(t.text or "")
    return "".join(parts).strip()


def paragraph_style_id(p):
    ppr = p.find("w:pPr", NS)
    if ppr is None:
        return None
    pstyle = ppr.find("w:pStyle", NS)
    if pstyle is None:
        return None
    return pstyle.attrib.get(f"{{{NS['w']}}}val")


def is_heading1(style_id, style_names):
    if not style_id:
        return False
    if style_id.lower() == "heading1":
        return True
    style_name = style_names.get(style_id, "")
    return "heading 1" in style_name or "标题 1" in style_name or "一级标题" in style_name


def extract_refs_from_paragraph(p):
    refs = []
    for blip in p.findall(".//a:blip", NS):
        rid = blip.attrib.get(f"{{{NS['r']}}}embed")
        if rid:
            refs.append(rid)
    for ole in p.findall(".//o:OLEObject", NS):
        rid = ole.attrib.get(f"{{{NS['r']}}}id") or ole.attrib.get("id")
        if rid:
            refs.append(rid)
    return refs


def classify_target(target: str):
    low = target.lower()
    if low.endswith((".png", ".jpg", ".jpeg", ".emf", ".bmp", ".gif")):
        return "image"
    if low.endswith(".xlsx"):
        return "excel"
    if low.endswith(".pdf"):
        return "pdf"
    if low.endswith(".bin"):
        return "ole_bin"
    return "unknown"


def summarize_embedded_file(docx_zip: ZipFile, target: str):
    full_name = f"word/{target}" if not target.startswith("word/") else target
    if full_name not in docx_zip.namelist():
        return {"readable": False, "reason": "embedded target not found"}

    data = docx_zip.read(full_name)
    summary = {
        "readable": True,
        "size_bytes": len(data),
    }

    kind = classify_target(target)
    if kind == "excel":
        try:
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
            ws = wb.active
            preview_rows = []
            for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
                preview_rows.append([c for c in row])
            summary["excel"] = {
                "sheet": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "preview": preview_rows,
            }
        except Exception as e:
            summary["readable"] = False
            summary["reason"] = f"excel parse failed: {e}"
    elif kind == "ole_bin":
        if b"%PDF" in data[:2048] or b"%PDF" in data:
            summary["possible_content"] = "pdf"
        else:
            summary["possible_content"] = "ole_binary_or_other"
    return summary


def parse_sample_docx(docx_path: Path, control_points: dict):
    sections = []
    current = None

    with ZipFile(docx_path) as docx_zip:
        rels = parse_relationships(docx_zip)
        style_names = parse_style_names(docx_zip)
        doc_root = ET.fromstring(docx_zip.read("word/document.xml"))
        body = doc_root.find("w:body", NS)

        for block in list(body):
            tag = block.tag.rsplit("}", 1)[-1]

            if tag == "p":
                text = paragraph_text(block)
                style_id = paragraph_style_id(block)

                if is_heading1(style_id, style_names):
                    heading_code = text.strip()
                    current = {
                        "heading": heading_code,
                        "control_point_match": control_points.get(heading_code),
                        "content": [],
                    }
                    sections.append(current)
                    continue

                if current is None:
                    continue

                refs = extract_refs_from_paragraph(block)
                entry = {
                    "type": "paragraph",
                    "text": text,
                    "embedded_items": [],
                }

                for rid in refs:
                    rel = rels.get(rid, {})
                    target = rel.get("target", "")
                    item = {
                        "rid": rid,
                        "target": target,
                        "kind": classify_target(target),
                        "summary": summarize_embedded_file(docx_zip, target),
                    }
                    entry["embedded_items"].append(item)

                if text or entry["embedded_items"]:
                    current["content"].append(entry)

            elif tag == "tbl":
                if current is None:
                    continue
                rows = []
                for tr in block.findall(".//w:tr", NS):
                    row_cells = []
                    for tc in tr.findall(".//w:tc", NS):
                        cell_text = "".join((t.text or "") for t in tc.findall(".//w:t", NS)).strip()
                        row_cells.append(cell_text)
                    if row_cells:
                        rows.append(row_cells)
                current["content"].append({"type": "table", "rows": rows})

    return sections


def main():
    if not DOCX_PATH.exists():
        raise FileNotFoundError(f"找不到文件: {DOCX_PATH}")
    if not CONTROL_XLSX_PATH.exists():
        raise FileNotFoundError(f"找不到文件: {CONTROL_XLSX_PATH}")

    control_points = load_control_points(CONTROL_XLSX_PATH)
    sections = parse_sample_docx(DOCX_PATH, control_points)

    report = {
        "docx_path": str(DOCX_PATH),
        "control_file": str(CONTROL_XLSX_PATH),
        "total_sections": len(sections),
        "matched_sections": sum(1 for s in sections if s.get("control_point_match")),
        "unmatched_sections": [s["heading"] for s in sections if not s.get("control_point_match")],
        "sections": sections,
    }

    OUTPUT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    print(f"[+] 可读性报告已生成: {OUTPUT_PATH}")
    print(f"[+] 一级标题数量: {report['total_sections']}")
    print(f"[+] 匹配到控制点: {report['matched_sections']}")
    if report["unmatched_sections"]:
        print(f"[!] 未匹配编号: {', '.join(report['unmatched_sections'])}")
    else:
        print("[+] 所有一级标题均匹配到控制点。")


if __name__ == "__main__":
    main()
